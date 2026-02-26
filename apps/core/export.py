"""Funciones para exportar datos a CSV y Excel."""

import csv
from io import BytesIO
from typing import Any

from django.http import HttpResponse

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def export_to_csv(
    data: list[dict[str, Any]], filename: str, headers: list[str] = None
) -> HttpResponse:
    """
    Exporta datos a un archivo CSV.

    Args:
        :param data: Lista de diccionarios con los datos
        :type data: List[Dict[str, Any]]
        :param filename: Nombre del archivo (sin extensión)
        :type filename: str
        :param headers: Lista de headers (opcional, usa keys del primer dict)
        :type headers: List[str]
    Returns:
        HttpResponse con el archivo CSV
    """
    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = f'attachment; filename="{filename}.csv"'
    response.write("\ufeff")  # BOM para Excel

    if not data:
        return response

    if headers is None:
        headers = list(data[0].keys())

    writer = csv.DictWriter(response, fieldnames=headers, extrasaction="ignore")
    writer.writeheader()
    writer.writerows(data)

    return response


def export_to_excel(
    data: list[dict[str, Any]], filename: str, headers: list[str] = None, sheet_name: str = "Datos"
) -> HttpResponse:
    """
    Exporta datos a un archivo Excel (.xlsx).

    Args:
        data: Lista de diccionarios con los datos
        filename: Nombre del archivo (sin extensión)
        headers: Lista de headers (opcional, usa keys del primer dict)
        sheet_name: Nombre de la hoja

    Returns:
        HttpResponse con el archivo Excel
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    if not data:
        # Archivo vacío
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}.xlsx"'
        return response

    if headers is None:
        headers = list(data[0].keys())

    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1a5276", end_color="1a5276", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Escribir headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Escribir datos
    for row_num, row_data in enumerate(data, 2):
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=row_num, column=col_num, value=row_data.get(header, ""))
            cell.border = thin_border

    # Ajustar ancho de columnas
    for col_num, header in enumerate(headers, 1):
        column_letter = get_column_letter(col_num)
        max_length = len(str(header))

        for row in ws.iter_rows(min_row=2, min_col=col_num, max_col=col_num):
            for cell in row:
                try:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
                except Exception:
                    # Si algo falla al calcular la longitud, ignoramos esa celda
                    continue

        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Guardar en buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}.xlsx"'
    return response


def queryset_to_dict_list(queryset, fields: list[str]) -> list[dict[str, Any]]:
    """
    Convierte un queryset a una lista de diccionarios.

    Args:
        queryset: QuerySet de Django
        fields: Lista de campos a incluir

    Returns:
        Lista de diccionarios
    """
    return list(queryset.values(*fields))
