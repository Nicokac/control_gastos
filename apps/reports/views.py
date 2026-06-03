"""
Vistas para dashboard y reportes.
"""

from datetime import datetime, time
from decimal import Decimal

from django.contrib.auth.mixins import LoginRequiredMixin
from django.db.models import Count, Q, Sum
from django.http import HttpResponse
from django.utils import timezone
from django.views.generic import TemplateView

from apps.core.utils import get_financial_period, get_month_date_range_exclusive, get_month_name
from apps.expenses.models import Expense
from apps.income.models import Income
from apps.savings.models import MovementType, Saving, SavingMovement, SavingStatus


class DashboardView(LoginRequiredMixin, TemplateView):
    """Dashboard principal con resumen financiero."""

    template_name = "reports/dashboard.html"

    def get_context_data(self, **kwargs):
        """Construye el contexto con todos los datos del dashboard."""
        context = super().get_context_data(**kwargs)

        user = self.request.user
        today = timezone.localdate()

        # Período seleccionado (por defecto: mes actual)
        try:
            selected_month = int(self.request.GET.get("month", today.month))
            selected_year = int(self.request.GET.get("year", today.year))
        except (ValueError, TypeError):
            selected_month = today.month
            selected_year = today.year

        # Limitar a rango válido: no futuro, no antes de 2020
        if selected_year > today.year or (
            selected_year == today.year and selected_month > today.month
        ):
            selected_month, selected_year = today.month, today.year
        selected_month = max(1, min(selected_month, 12))
        selected_year = max(2020, selected_year)

        is_current_period = selected_month == today.month and selected_year == today.year

        # Mes anterior y siguiente para navegación
        if selected_month == 1:
            prev_month, prev_nav_year = 12, selected_year - 1
        else:
            prev_month, prev_nav_year = selected_month - 1, selected_year

        if selected_month == 12:
            next_month, next_nav_year = 1, selected_year + 1
        else:
            next_month, next_nav_year = selected_month + 1, selected_year

        # Años disponibles para el gráfico de evolución
        first_year = self._get_first_data_year(user, today.year)
        year_choices = list(range(first_year, today.year + 1))

        # Para el gráfico de evolución: usar el año seleccionado
        evolution_month = selected_month if selected_year == today.year else 12

        context["current_month"] = selected_month
        context["current_year"] = selected_year
        context["month_name"] = get_month_name(selected_month)
        context["today"] = today
        context["is_current_period"] = is_current_period
        context["prev_month"] = prev_month
        context["prev_nav_year"] = prev_nav_year
        context["next_month"] = next_month
        context["next_nav_year"] = next_nav_year
        context["selected_year"] = selected_year
        context["year_choices"] = year_choices
        context["alert_threshold"] = user.alert_threshold

        # Período financiero del usuario
        start_day = getattr(user, "financial_month_start_day", 1)
        fin_start, fin_end = get_financial_period(selected_month, selected_year, start_day)
        context["financial_period_start"] = fin_start
        context["financial_period_end"] = fin_end
        context["financial_start_day"] = start_day

        # Días transcurridos y totales del período financiero (para "Día X de Y")
        period_total_days = (fin_end - fin_start).days
        if is_current_period and fin_start <= today < fin_end:
            period_day = (today - fin_start).days + 1
        else:
            period_day = period_total_days
        context["period_day"] = period_day
        context["period_total_days"] = period_total_days

        # Obtener datos de cada módulo
        context.update(
            self._get_balance_data(user, selected_month, selected_year, fin_start, fin_end)
        )
        context.update(self._get_savings_data(user))
        context.update(self._get_recurring_data(user, selected_month, selected_year))
        context.update(self._get_recurring_income_data(user, selected_month, selected_year))
        context.update(self._get_recent_transactions(user))
        context.update(self._get_expense_distribution(user, selected_month, selected_year))
        context.update(self._get_monthly_evolution(user, evolution_month, selected_year))

        return context

    def _get_first_data_year(self, user, current_year):
        """Retorna el año más antiguo con datos del usuario (mínimo 2020)."""
        expense_year = (
            Expense.objects.filter(user=user)
            .order_by("date")
            .values_list("date__year", flat=True)
            .first()
        )
        income_year = (
            Income.objects.filter(user=user)
            .order_by("date")
            .values_list("date__year", flat=True)
            .first()
        )
        candidates = [y for y in [expense_year, income_year] if y]
        return min(candidates) if candidates else current_year

    def _get_balance_data(self, user, month, year, cur_start=None, cur_end=None):
        """
        Obtiene datos de balance: ingresos vs gastos.

        Optimizado: 2 queries en lugar de 6.
        Usa el período financiero del usuario si se provee cur_start/cur_end.
        """

        # Rango mes actual (financiero o calendario)
        if cur_start is None or cur_end is None:
            cur_start, cur_end = get_month_date_range_exclusive(month, year)

        # Mes anterior: mismo start_day pero mes previo
        if month == 1:
            prev_month, prev_year = 12, year - 1
        else:
            prev_month, prev_year = month - 1, year

        # Para el mes anterior usamos el mismo start_day para consistencia
        start_day = cur_start.day if cur_start.day != 1 else 1
        from apps.core.utils import get_financial_period as _gfp

        prev_start, prev_end = _gfp(prev_month, prev_year, start_day)

        # Query 1: Gastos (mes actual + anterior)
        expense_data = (
            Expense.objects.filter(user=user)
            .filter(
                Q(date__gte=cur_start, date__lt=cur_end)
                | Q(date__gte=prev_start, date__lt=prev_end)
            )
            .aggregate(
                current=Sum("amount_ars", filter=Q(date__gte=cur_start, date__lt=cur_end)),
                previous=Sum("amount_ars", filter=Q(date__gte=prev_start, date__lt=prev_end)),
            )
        )

        # Query 2: Ingresos (mes actual + anterior)
        income_data = (
            Income.objects.filter(user=user)
            .filter(
                Q(date__gte=cur_start, date__lt=cur_end)
                | Q(date__gte=prev_start, date__lt=prev_end)
            )
            .aggregate(
                current=Sum("amount_ars", filter=Q(date__gte=cur_start, date__lt=cur_end)),
                previous=Sum("amount_ars", filter=Q(date__gte=prev_start, date__lt=prev_end)),
            )
        )

        # Extraer valores
        expense_total = expense_data["current"] or Decimal("0")
        prev_expense_total = expense_data["previous"] or Decimal("0")
        income_total = income_data["current"] or Decimal("0")
        prev_income_total = income_data["previous"] or Decimal("0")

        # Query 3: Depósitos a ahorro del mes actual
        savings_deposits_data = SavingMovement.objects.filter(
            saving__user=user,
            type=MovementType.DEPOSIT,
            date__gte=cur_start,
            date__lt=cur_end,
        ).aggregate(total=Sum("amount"))
        savings_deposits_total = savings_deposits_data["total"] or Decimal("0")

        # Balance (ingresos - gastos - ahorro)
        balance = income_total - expense_total - savings_deposits_total

        # Porcentaje de gastos vs ingresos
        if income_total > 0:
            expense_percentage = round((expense_total / income_total) * 100, 1)
        else:
            expense_percentage = 0

        # Variación porcentual de gastos
        if prev_expense_total > 0:
            expense_variation = round(
                ((expense_total - prev_expense_total) / prev_expense_total) * 100, 1
            )
        else:
            expense_variation = 0

        # Variación porcentual de ingresos
        if prev_income_total > 0:
            income_variation = round(
                ((income_total - prev_income_total) / prev_income_total) * 100, 1
            )
        else:
            income_variation = 0

        return {
            "income_total": income_total,
            "expense_total": expense_total,
            "savings_deposits_total": savings_deposits_total,
            "balance": balance,
            "balance_is_positive": balance >= 0,
            "expense_percentage": expense_percentage,
            "expense_variation": expense_variation,
            "income_variation": income_variation,
            "prev_income_total": prev_income_total,
            "prev_month_name": get_month_name(prev_month),
        }

    def _get_savings_data(self, user):
        """Obtiene datos de metas de ahorro activas."""
        today = timezone.localdate()
        month_start_date, month_end_date = get_month_date_range_exclusive(today.month, today.year)

        # Convertir a datetime aware para filtrar DateTimeField (updated_at)
        month_start = timezone.make_aware(datetime.combine(month_start_date, time.min))
        month_end = timezone.make_aware(datetime.combine(month_end_date, time.min))

        # Query 1: Aggregates (totales + count completed) en una sola query
        aggregates = Saving.objects.filter(user=user).aggregate(
            total_target=Sum(
                "target_amount", filter=Q(status=SavingStatus.ACTIVE), default=Decimal("0")
            ),
            total_current=Sum(
                "current_amount", filter=Q(status=SavingStatus.ACTIVE), default=Decimal("0")
            ),
            active_count=Count("pk", filter=Q(status=SavingStatus.ACTIVE)),
            completed_this_month=Count(
                "pk",
                filter=Q(
                    status=SavingStatus.COMPLETED,
                    updated_at__gte=month_start,
                    updated_at__lt=month_end,
                ),
            ),
        )

        total_target = aggregates["total_target"]
        total_current = aggregates["total_current"]

        if total_target > 0:
            overall_progress = round((total_current / total_target) * 100, 1)
        else:
            overall_progress = 0

        from django.db.models import DecimalField, ExpressionWrapper, F

        # Query 2: Top 3 metas activas por progreso (ordenado en DB)
        progress_expr = ExpressionWrapper(
            (F("current_amount") * 100) / F("target_amount"),
            output_field=DecimalField(max_digits=7, decimal_places=2),
        )

        top_savings = list(
            Saving.objects.filter(user=user, status=SavingStatus.ACTIVE)
            .annotate(progress_db=progress_expr)
            .order_by("-progress_db", "-updated_at")[:3]
        )

        return {
            "savings_count": aggregates["active_count"],
            "savings_total_target": total_target,
            "savings_total_current": total_current,
            "savings_progress": overall_progress,
            "savings_completed_month": aggregates["completed_this_month"],
            "top_savings": top_savings,
        }

    def _get_recurring_data(self, user, month, year):
        """Obtiene el estado de gastos fijos del mes actual."""
        from apps.recurring.models import RecurringExpense

        recurrents = RecurringExpense.objects.filter(user=user, is_active=True).prefetch_related(
            "expenses"
        )
        total_active = len(recurrents)
        total_paid = sum(1 for r in recurrents if r.is_paid_in(month, year))
        overdue = sum(
            1
            for r in recurrents
            if not r.is_paid_in(month, year) and r.status_for(month, year) == "overdue"
        )
        pending = [
            {"rec": r, "overdue": r.status_for(month, year) == "overdue"}
            for r in recurrents
            if not r.is_paid_in(month, year)
        ]

        return {
            "recurring_total": total_active,
            "recurring_paid": total_paid,
            "recurring_overdue": overdue,
            "recurring_pending": pending,
        }

    def _get_recurring_income_data(self, user, month, year):
        """Obtiene el estado de ingresos fijos del mes actual."""
        from apps.recurring_income.models import RecurringIncome

        recurrents = RecurringIncome.objects.filter(user=user, is_active=True).prefetch_related(
            "incomes"
        )
        total_active = len(recurrents)
        total_collected = sum(1 for r in recurrents if r.is_collected_in(month, year))
        overdue = sum(
            1
            for r in recurrents
            if not r.is_collected_in(month, year) and r.status_for(month, year) == "overdue"
        )
        pending = [
            {"rec": r, "overdue": r.status_for(month, year) == "overdue"}
            for r in recurrents
            if not r.is_collected_in(month, year)
        ]

        return {
            "recurring_income_total": total_active,
            "recurring_income_collected": total_collected,
            "recurring_income_overdue": overdue,
            "recurring_income_pending": pending,
        }

    def _get_recent_transactions(self, user):
        """
        Obtiene las últimas 5 transacciones (gastos e ingresos combinados).

        Optimizado: Usa queries separadas pero obtiene suficientes registros
        para garantizar las 5 transacciones más recientes reales.
        """
        # Obtener últimos 5 de cada tipo para garantizar cobertura
        recent_expenses = list(
            Expense.objects.filter(user=user)
            .select_related("category")
            .order_by("-date", "-created_at")[:5]
        )

        recent_incomes = list(
            Income.objects.filter(user=user)
            .select_related("category")
            .order_by("-date", "-created_at")[:5]
        )

        # Combinar en lista unificada
        transactions = []

        for expense in recent_expenses:
            transactions.append(
                {
                    "type": "expense",
                    "date": expense.date,
                    "created_at": expense.created_at,
                    "description": expense.description,
                    "amount": expense.amount_ars,
                    "formatted_amount": expense.formatted_amount,
                    "category": expense.category,
                    "pk": expense.pk,
                }
            )

        for income in recent_incomes:
            transactions.append(
                {
                    "type": "income",
                    "date": income.date,
                    "created_at": income.created_at,
                    "description": income.description,
                    "amount": income.amount_ars,
                    "formatted_amount": income.formatted_amount,
                    "category": income.category,
                    "pk": income.pk,
                }
            )

        # Ordenar por fecha y created_at descendente, luego cortar a 5
        transactions.sort(key=lambda x: (x["date"], x["created_at"]), reverse=True)

        return {
            "recent_transactions": transactions[:5],
        }

    def _get_monthly_evolution(self, user, month, year):
        """
        Retorna ingresos, gastos y ahorro acumulados por mes,
        desde enero del año actual hasta el mes actual inclusive.
        """
        months = range(1, month + 1)
        labels = []
        income_data = []
        expense_data = []
        savings_data = []
        balance_data = []

        # Rango completo: enero → fin del mes actual (1 query por modelo)
        year_start, _ = get_month_date_range_exclusive(1, year)
        _, period_end = get_month_date_range_exclusive(month, year)

        expenses_qs = (
            Expense.objects.filter(user=user, date__gte=year_start, date__lt=period_end)
            .values("date__month")
            .annotate(total=Sum("amount_ars"))
        )
        incomes_qs = (
            Income.objects.filter(user=user, date__gte=year_start, date__lt=period_end)
            .values("date__month")
            .annotate(total=Sum("amount_ars"))
        )
        savings_qs = (
            SavingMovement.objects.filter(
                saving__user=user,
                type=MovementType.DEPOSIT,
                date__gte=year_start,
                date__lt=period_end,
            )
            .values("date__month")
            .annotate(total=Sum("amount"))
        )

        expense_by_month = {row["date__month"]: float(row["total"]) for row in expenses_qs}
        income_by_month = {row["date__month"]: float(row["total"]) for row in incomes_qs}
        savings_by_month = {row["date__month"]: float(row["total"]) for row in savings_qs}

        for m in months:
            inc = income_by_month.get(m, 0)
            exp = expense_by_month.get(m, 0)
            sav = savings_by_month.get(m, 0)
            labels.append(get_month_name(m))
            income_data.append(inc)
            expense_data.append(exp)
            savings_data.append(sav)
            balance_data.append(round(inc - exp - sav, 2))

        month_urls = [f"?month={m}&year={year}" for m in months]

        return {
            "evolution_labels": labels,
            "evolution_income": income_data,
            "evolution_expenses": expense_data,
            "evolution_savings": savings_data,
            "evolution_balance": balance_data,
            "evolution_month_urls": month_urls,
        }

    def _get_expense_distribution(self, user, month, year):
        """
        Obtiene distribución de gastos agrupada por grupo padre de la categoría.

        Retorna:
        - ranking_distribution: grupos ordenados por total (para el ranking con scroll)
        - chart_*: top 5 grupos + "Otros" agrupado (para el donut)
        """
        start_date, end_date = get_month_date_range_exclusive(month, year)

        raw = list(
            Expense.objects.filter(user=user, date__gte=start_date, date__lt=end_date)
            .select_related("category__parent")
            .values(
                "category__parent__id",
                "category__parent__name",
                "category__parent__color",
                "category__id",
                "category__name",
                "category__color",
                "category__icon",
            )
            .annotate(total=Sum("amount_ars"))
            .order_by("-total")
        )

        # Agrupar por grupo padre; si no tiene padre usar la categoría misma como grupo
        groups: dict = {}
        for item in raw:
            group_name = (
                item["category__parent__name"] or item["category__name"] or "Sin clasificar"
            )
            group_color = item["category__parent__color"] or item["category__color"] or "#6c757d"
            group_pk = item["category__parent__id"] or item["category__id"]
            if group_name not in groups:
                groups[group_name] = {
                    "name": group_name,
                    "color": group_color,
                    "pk": group_pk,
                    "total": Decimal("0"),
                    "subcategories": [],
                }
            groups[group_name]["total"] += item["total"]
            # Solo agregar subcategoría si la categoría tiene padre (no es el grupo mismo)
            if item["category__parent__name"]:
                groups[group_name]["subcategories"].append(
                    {
                        "name": item["category__name"],
                        "pk": item["category__id"],
                        "icon": item["category__icon"] or "",
                        "total": item["total"],
                    }
                )

        grand_total = sum(float(g["total"]) for g in groups.values())

        ranking = sorted(groups.values(), key=lambda g: g["total"], reverse=True)
        for item in ranking:
            item["percentage"] = (
                round((float(item["total"]) / grand_total) * 100, 1) if grand_total > 0 else 0
            )

        # Donut: top 5 grupos + "Otros"
        TOP_N = 5
        top = ranking[:TOP_N]
        others = ranking[TOP_N:]

        chart_labels = [r["name"] for r in top]
        chart_data = [float(r["total"]) for r in top]
        chart_colors = [r["color"] for r in top]

        if others:
            others_total = sum(float(r["total"]) for r in others)
            chart_labels.append("Otros")
            chart_data.append(others_total)
            chart_colors.append("#6c757d")

        top_category = ranking[0] if ranking else None

        return {
            "ranking_distribution": ranking,
            "chart_labels": chart_labels,
            "chart_data": chart_data,
            "chart_colors": chart_colors,
            "chart_grand_total": grand_total,
            "top_category": top_category,
            "expense_distribution": ranking,
        }


class DashboardExportView(DashboardView):
    """Exporta el resumen del período activo como .xlsx."""

    def get(self, request, *args, **kwargs):
        import contextlib
        import io

        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill

        user = request.user
        today = timezone.localdate()

        with contextlib.suppress(ValueError, TypeError):
            selected_month = int(request.GET.get("month", today.month))
            selected_year = int(request.GET.get("year", today.year))

        if "selected_month" not in dir():
            selected_month, selected_year = today.month, today.year

        month_name = get_month_name(selected_month)
        start_day = getattr(user, "financial_month_start_day", 1)
        fin_start, fin_end = get_financial_period(selected_month, selected_year, start_day)

        # Datos
        balance = self._get_balance_data(user, selected_month, selected_year, fin_start, fin_end)
        distribution = self._get_expense_distribution(user, selected_month, selected_year)

        income_by_cat = (
            Income.objects.filter(user=user, date__gte=fin_start, date__lt=fin_end)
            .select_related("category__parent")
            .values("category__parent__name", "category__name")
            .annotate(total=Sum("amount_ars"))
            .order_by("-total")
        )

        # Estilos
        wb = Workbook()
        ws = wb.active
        ws.title = f"{month_name} {selected_year}"

        dark_blue = "1F4E79"
        mid_blue = "2E75B6"
        light_blue = "D6E4F0"
        light_gray = "F2F2F2"
        green_fill = "E2EFDA"

        def header_style(cell, bg=dark_blue):
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.alignment = Alignment(horizontal="center")

        def section_style(cell, bg=mid_blue):
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor=bg)

        def subtotal_style(cell, bg=light_blue):
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor=bg)

        right = Alignment(horizontal="right")
        center = Alignment(horizontal="center")

        row = 1

        # ── Título ──────────────────────────────────────────────────────────
        title = ws.cell(
            row=row, column=1, value=f"Resumen financiero — {month_name} {selected_year}"
        )
        title.font = Font(bold=True, size=13)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        row += 2

        # ── Sección 1: Balance del período ──────────────────────────────────
        sec = ws.cell(row=row, column=1, value="Balance del período")
        section_style(sec)
        ws.cell(row=row, column=2).fill = PatternFill("solid", fgColor=mid_blue)
        ws.cell(row=row, column=3).fill = PatternFill("solid", fgColor=mid_blue)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        row += 1

        balance_rows = [
            ("Ingresos", balance["income_total"]),
            ("Gastos", balance["expense_total"]),
            ("Ahorro depositado", balance["savings_deposits_total"]),
            ("Balance", balance["balance"]),
            ("% gastado sobre ingresos", f"{balance['expense_percentage']}%"),
        ]
        for label, value in balance_rows:
            ws.cell(row=row, column=1, value=label)
            val_cell = ws.cell(
                row=row, column=2, value=float(value) if not isinstance(value, str) else value
            )
            val_cell.alignment = right
            if label == "Balance":
                subtotal_style(ws.cell(row=row, column=1), bg=light_blue)
                subtotal_style(val_cell, bg=light_blue)
            elif label == "% gastado sobre ingresos":
                ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor=light_gray)
                val_cell.fill = PatternFill("solid", fgColor=light_gray)
            row += 1

        row += 1

        # ── Sección 2: Gastos por categoría ─────────────────────────────────
        sec = ws.cell(row=row, column=1, value="Gastos por categoría")
        section_style(sec)
        ws.cell(row=row, column=2).fill = PatternFill("solid", fgColor=mid_blue)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        row += 1

        h1 = ws.cell(row=row, column=1, value="Grupo")
        h2 = ws.cell(row=row, column=2, value="Monto")
        h3 = ws.cell(row=row, column=3, value="%")
        for h in [h1, h2, h3]:
            header_style(h, bg="404040")
        row += 1

        for group in distribution["ranking_distribution"]:
            ws.cell(row=row, column=1, value=group["name"])
            ws.cell(row=row, column=2, value=float(group["total"])).alignment = right
            ws.cell(row=row, column=3, value=f"{group['percentage']}%").alignment = center
            row += 1

        # Subtotal gastos
        sub = ws.cell(row=row, column=1, value="Total gastos")
        subtotal_style(sub, light_blue)
        sub_val = ws.cell(row=row, column=2, value=float(balance["expense_total"]))
        subtotal_style(sub_val, light_blue)
        sub_val.alignment = right
        ws.cell(row=row, column=3).fill = PatternFill("solid", fgColor=light_blue)
        row += 2

        # ── Sección 3: Ingresos por categoría ───────────────────────────────
        sec = ws.cell(row=row, column=1, value="Ingresos por categoría")
        section_style(sec, bg="375623")
        ws.cell(row=row, column=2).fill = PatternFill("solid", fgColor="375623")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        row += 1

        h1 = ws.cell(row=row, column=1, value="Categoría")
        h2 = ws.cell(row=row, column=2, value="Monto")
        for h in [h1, h2]:
            header_style(h, bg="404040")
        ws.cell(row=row, column=3).fill = PatternFill("solid", fgColor="404040")
        row += 1

        for inc_row in income_by_cat:
            cat_name = inc_row["category__parent__name"] or inc_row["category__name"]
            ws.cell(row=row, column=1, value=cat_name)
            ws.cell(row=row, column=2, value=float(inc_row["total"])).alignment = right
            row += 1

        sub = ws.cell(row=row, column=1, value="Total ingresos")
        sub.font = Font(bold=True)
        sub.fill = PatternFill("solid", fgColor=green_fill)
        sub_val = ws.cell(row=row, column=2, value=float(balance["income_total"]))
        sub_val.font = Font(bold=True)
        sub_val.fill = PatternFill("solid", fgColor=green_fill)
        sub_val.alignment = right

        # Anchos de columna
        ws.column_dimensions["A"].width = 32
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 10

        # Respuesta
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"resumen_{month_name.lower()}_{selected_year}.xlsx"
        response = HttpResponse(
            output.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response


def _build_annual_data(user, year):
    """
    Construye los datos del reporte anual para un año dado.
    Retorna lista de 12 dicts con totales por mes + totales anuales.
    3 queries totales independientemente del año.
    """
    from decimal import Decimal as D

    year_start, year_end = get_month_date_range_exclusive(1, year)
    _, year_end = get_month_date_range_exclusive(12, year)

    expenses_qs = (
        Expense.objects.filter(user=user, date__year=year)
        .values("date__month")
        .annotate(total=Sum("amount_ars"))
    )
    incomes_qs = (
        Income.objects.filter(user=user, date__year=year)
        .values("date__month")
        .annotate(total=Sum("amount_ars"))
    )
    savings_qs = (
        SavingMovement.objects.filter(saving__user=user, type=MovementType.DEPOSIT, date__year=year)
        .values("date__month")
        .annotate(total=Sum("amount"))
    )

    expense_by_month = {row["date__month"]: D(str(row["total"])) for row in expenses_qs}
    income_by_month = {row["date__month"]: D(str(row["total"])) for row in incomes_qs}
    savings_by_month = {row["date__month"]: D(str(row["total"])) for row in savings_qs}

    rows = []
    for m in range(1, 13):
        inc = income_by_month.get(m, D("0"))
        exp = expense_by_month.get(m, D("0"))
        sav = savings_by_month.get(m, D("0"))
        bal = inc - exp - sav
        rate = round((exp / inc * 100), 1) if inc > 0 else D("0")
        rows.append(
            {
                "month": m,
                "month_name": get_month_name(m),
                "income": inc,
                "expense": exp,
                "savings": sav,
                "balance": bal,
                "expense_rate": rate,
                "has_data": inc > 0 or exp > 0 or sav > 0,
            }
        )

    total_income = sum(r["income"] for r in rows)
    total_expense = sum(r["expense"] for r in rows)
    total_savings = sum(r["savings"] for r in rows)
    total_balance = total_income - total_expense - total_savings
    total_rate = round((total_expense / total_income * 100), 1) if total_income > 0 else D("0")

    totals = {
        "income": total_income,
        "expense": total_expense,
        "savings": total_savings,
        "balance": total_balance,
        "expense_rate": total_rate,
    }
    return rows, totals


class AnnualReportView(LoginRequiredMixin, TemplateView):
    """Reporte anual — tabla comparativa mes a mes."""

    template_name = "reports/annual_report.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user = self.request.user
        today = timezone.localdate()

        try:
            year = int(self.request.GET.get("year", today.year))
        except (ValueError, TypeError):
            year = today.year

        year = max(2020, min(year, today.year))

        first_year = (
            Expense.objects.filter(user=user)
            .order_by("date")
            .values_list("date__year", flat=True)
            .first()
            or today.year
        )
        year_choices = list(range(min(first_year, today.year), today.year + 1))

        rows, totals = _build_annual_data(user, year)

        context["year"] = year
        context["year_choices"] = year_choices
        context["rows"] = rows
        context["totals"] = totals
        context["today"] = today
        return context


class AnnualReportExportView(LoginRequiredMixin, TemplateView):
    """Exporta el reporte anual como .xlsx."""

    def get(self, request, *args, **kwargs):
        import contextlib
        import io

        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill

        user = request.user
        today = timezone.localdate()

        with contextlib.suppress(ValueError, TypeError):
            year = int(request.GET.get("year", today.year))
        if "year" not in dir():
            year = today.year

        rows, totals = _build_annual_data(user, year)

        wb = Workbook()
        ws = wb.active
        ws.title = str(year)

        dark_blue = "1F4E79"
        light_gray = "F2F2F2"
        green = "E2EFDA"
        red = "FCE4D6"

        def hdr(cell):
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor=dark_blue)
            cell.alignment = Alignment(horizontal="center")

        right = Alignment(horizontal="right")
        center = Alignment(horizontal="center")

        # Título
        t = ws.cell(row=1, column=1, value=f"Reporte anual {year}")
        t.font = Font(bold=True, size=13)
        ws.merge_cells("A1:F1")
        ws.row_dimensions[1].height = 20

        # Header de columnas
        headers = ["Mes", "Ingresos", "Gastos", "Ahorro", "Balance", "% Gastado"]
        for col, h in enumerate(headers, start=1):
            hdr(ws.cell(row=2, column=col, value=h))

        # Filas de datos
        for i, row in enumerate(rows, start=3):
            ws.cell(row=i, column=1, value=row["month_name"])
            ws.cell(row=i, column=2, value=float(row["income"])).alignment = right
            ws.cell(row=i, column=3, value=float(row["expense"])).alignment = right
            ws.cell(row=i, column=4, value=float(row["savings"])).alignment = right

            bal_cell = ws.cell(row=i, column=5, value=float(row["balance"]))
            bal_cell.alignment = right
            if row["balance"] < 0:
                bal_cell.fill = PatternFill("solid", fgColor=red)
            elif row["balance"] > 0:
                bal_cell.fill = PatternFill("solid", fgColor=green)

            ws.cell(row=i, column=6, value=f"{row['expense_rate']}%").alignment = center

            if not row["has_data"]:
                for col in range(1, 7):
                    ws.cell(row=i, column=col).fill = PatternFill("solid", fgColor=light_gray)

        # Fila de totales
        total_row = 15
        for col in range(1, 7):
            ws.cell(row=total_row, column=col).fill = PatternFill("solid", fgColor=dark_blue)
            ws.cell(row=total_row, column=col).font = Font(bold=True, color="FFFFFF")

        ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True, color="FFFFFF")
        ws.cell(row=total_row, column=2, value=float(totals["income"])).alignment = right
        ws.cell(row=total_row, column=3, value=float(totals["expense"])).alignment = right
        ws.cell(row=total_row, column=4, value=float(totals["savings"])).alignment = right
        ws.cell(row=total_row, column=5, value=float(totals["balance"])).alignment = right
        ws.cell(row=total_row, column=6, value=f"{totals['expense_rate']}%").alignment = center
        for col in range(1, 7):
            ws.cell(row=total_row, column=col).font = Font(bold=True, color="FFFFFF")

        # Anchos
        ws.column_dimensions["A"].width = 14
        for col in ["B", "C", "D", "E"]:
            ws.column_dimensions[col].width = 18
        ws.column_dimensions["F"].width = 12

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="reporte_anual_{year}.xlsx"'
        return response
