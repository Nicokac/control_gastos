"""Vistas para gastos compartidos del hogar."""

from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db.models import Sum
from django.http import HttpResponse
from django.urls import reverse_lazy
from django.utils import timezone
from django.views.generic import CreateView, DeleteView, ListView, UpdateView, View

from apps.core.utils import get_month_name, get_months_choices, get_years_choices
from apps.core.views import UserFormKwargsMixin

from .forms import HouseholdMemberForm, SharedExpenseForm
from .models import HouseholdMember, SharedExpense


class SharedExpenseListView(LoginRequiredMixin, ListView):
    model = SharedExpense
    template_name = "shared_expenses/shared_expense_list.html"
    context_object_name = "expenses"
    paginate_by = 50

    def get_queryset(self):
        qs = SharedExpense.objects.filter(user=self.request.user).select_related(
            "category", "category__parent", "paid_by"
        )
        month = self.request.GET.get("month")
        year = self.request.GET.get("year")
        if not (month and year):
            today = timezone.localdate()
            month = str(today.month)
            year = str(today.year)
        import contextlib

        with contextlib.suppress(ValueError, TypeError):
            qs = qs.filter(date__month=int(month), date__year=int(year))
        return qs.order_by("-date", "-created_at")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user = self.request.user
        today = timezone.localdate()

        month_str = self.request.GET.get("month", str(today.month))
        year_str = self.request.GET.get("year", str(today.year))
        try:
            month = int(month_str)
            year = int(year_str)
        except (ValueError, TypeError):
            month, year = today.month, today.year

        context["current_month"] = month
        context["current_year"] = year
        context["current_month_name"] = get_month_name(month)
        context["months"] = get_months_choices()
        context["years"] = [y for y, _ in get_years_choices()]

        qs = self.object_list
        members = list(HouseholdMember.objects.filter(user=user))
        context["members"] = members
        context["has_members"] = bool(members)

        # Cuánto gastó cada persona en el período
        owner_total = qs.filter(paid_by__isnull=True).aggregate(t=Sum("amount_ars"))["t"] or 0
        member_totals = [
            {
                "member": m,
                "total": qs.filter(paid_by=m).aggregate(t=Sum("amount_ars"))["t"] or 0,
            }
            for m in members
        ]
        grand_total = owner_total + sum(v["total"] for v in member_totals)

        context["owner_total"] = owner_total
        context["member_totals"] = member_totals
        context["grand_total"] = grand_total

        # Gastos agrupados por categoría — tabla estilo planilla con columnas por persona
        # Estructura: { group_pk: { group, expenses, subtotal_owner, subtotal_members: {member_pk: total} } }
        grouped = {}
        for expense in qs:
            group = expense.category.parent or expense.category
            key = group.pk
            if key not in grouped:
                grouped[key] = {
                    "group": group,
                    "expenses": [],
                    "subtotal_owner": 0,
                    "subtotal_members": {m.pk: 0 for m in members},
                    "subtotal_total": 0,
                }
            grouped[key]["expenses"].append(expense)
            grouped[key]["subtotal_total"] += expense.amount_ars
            if expense.paid_by_id:
                grouped[key]["subtotal_members"][expense.paid_by_id] = (
                    grouped[key]["subtotal_members"].get(expense.paid_by_id, 0) + expense.amount_ars
                )
            else:
                grouped[key]["subtotal_owner"] += expense.amount_ars

        # Convertir subtotal_members a lista ordenada por member para el template
        for g in grouped.values():
            g["subtotals_by_member"] = [
                {"member": m, "subtotal": g["subtotal_members"].get(m.pk, 0)} for m in members
            ]

        context["grouped_expenses"] = list(grouped.values())

        return context


class SharedExpenseCreateView(LoginRequiredMixin, UserFormKwargsMixin, CreateView):
    model = SharedExpense
    form_class = SharedExpenseForm
    template_name = "shared_expenses/shared_expense_form.html"
    success_url = reverse_lazy("shared_expenses:list")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["is_edit"] = False
        context["has_members"] = HouseholdMember.objects.filter(user=self.request.user).exists()
        return context

    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(
            self.request, f"Gasto '{self.object.description}' registrado correctamente."
        )
        return response

    def form_invalid(self, form):
        messages.error(self.request, "No pudimos guardar el gasto. Revisá los campos marcados.")
        return super().form_invalid(form)


class SharedExpenseUpdateView(LoginRequiredMixin, UserFormKwargsMixin, UpdateView):
    model = SharedExpense
    form_class = SharedExpenseForm
    template_name = "shared_expenses/shared_expense_form.html"
    success_url = reverse_lazy("shared_expenses:list")

    def get_queryset(self):
        return SharedExpense.objects.filter(user=self.request.user)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["is_edit"] = True
        context["has_members"] = HouseholdMember.objects.filter(user=self.request.user).exists()
        return context

    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(
            self.request, f"Gasto '{self.object.description}' actualizado correctamente."
        )
        return response

    def form_invalid(self, form):
        messages.error(self.request, "No pudimos guardar el gasto. Revisá los campos marcados.")
        return super().form_invalid(form)


class SharedExpenseDeleteView(LoginRequiredMixin, DeleteView):
    model = SharedExpense
    template_name = "shared_expenses/shared_expense_confirm_delete.html"
    success_url = reverse_lazy("shared_expenses:list")

    def get_queryset(self):
        return SharedExpense.objects.filter(user=self.request.user)

    def form_valid(self, form):
        name = self.get_object().description
        response = super().form_valid(form)
        messages.success(self.request, f"Gasto '{name}' eliminado correctamente.")
        return response


class HouseholdMemberListView(LoginRequiredMixin, ListView):
    model = HouseholdMember
    template_name = "shared_expenses/members.html"
    context_object_name = "members"

    def get_queryset(self):
        return HouseholdMember.objects.filter(user=self.request.user)


class HouseholdMemberCreateView(LoginRequiredMixin, UserFormKwargsMixin, CreateView):
    model = HouseholdMember
    form_class = HouseholdMemberForm
    template_name = "shared_expenses/member_form.html"
    success_url = reverse_lazy("shared_expenses:members")

    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(self.request, f"'{self.object.name}' agregado al hogar.")
        return response


class HouseholdMemberDeleteView(LoginRequiredMixin, DeleteView):
    model = HouseholdMember
    template_name = "shared_expenses/member_confirm_delete.html"
    success_url = reverse_lazy("shared_expenses:members")

    def get_queryset(self):
        return HouseholdMember.objects.filter(user=self.request.user)

    def form_valid(self, form):
        name = self.get_object().name
        response = super().form_valid(form)
        messages.success(self.request, f"'{name}' eliminado del hogar.")
        return response


class SharedExpenseExportView(LoginRequiredMixin, View):
    """Exporta los gastos compartidos del período como .xlsx estilo planilla."""

    def get(self, request):
        import io

        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        user = request.user
        today = timezone.localdate()

        try:
            month = int(request.GET.get("month", today.month))
            year = int(request.GET.get("year", today.year))
        except (ValueError, TypeError):
            month, year = today.month, today.year

        month_name = get_month_name(month)
        members = list(HouseholdMember.objects.filter(user=user))

        qs = (
            SharedExpense.objects.filter(user=user, date__month=month, date__year=year)
            .select_related("category", "category__parent", "paid_by")
            .order_by("category__parent__name", "category__name", "date")
        )

        # Agrupar por categoría
        grouped = {}
        for expense in qs:
            group = expense.category.parent or expense.category
            key = group.pk
            if key not in grouped:
                grouped[key] = {"group": group, "expenses": []}
            grouped[key]["expenses"].append(expense)

        wb = Workbook()
        ws = wb.active
        ws.title = f"{month_name} {year}"

        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="1F4E79")
        group_font = Font(bold=True)
        group_fill = PatternFill("solid", fgColor="D6E4F0")
        subtotal_fill = PatternFill("solid", fgColor="EBF5FB")
        total_fill = PatternFill("solid", fgColor="1F4E79")
        total_font = Font(bold=True, color="FFFFFF")
        center = Alignment(horizontal="center")
        right = Alignment(horizontal="right")

        # Encabezado de columnas
        col_headers = ["Descripción", "Fecha", "Yo"] + [m.name for m in members]
        for col, header in enumerate(col_headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center if col > 1 else Alignment()

        # Totales acumulados para la fila final
        owner_total = 0
        member_totals = {m.pk: 0 for m in members}

        row = 2
        for group_data in grouped.values():
            group = group_data["group"]
            group_subtotal_owner = 0
            group_subtotal_members = {m.pk: 0 for m in members}

            # Header de grupo
            cell = ws.cell(row=row, column=1, value=group.name.upper())
            cell.font = group_font
            cell.fill = group_fill
            for col in range(2, len(col_headers) + 1):
                ws.cell(row=row, column=col).fill = group_fill
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(col_headers))
            row += 1

            # Filas de gastos
            for expense in group_data["expenses"]:
                ws.cell(row=row, column=1, value=expense.description)
                ws.cell(
                    row=row, column=2, value=expense.date.strftime("%d/%m/%Y")
                ).alignment = center

                if not expense.paid_by:
                    ws.cell(row=row, column=3, value=float(expense.amount_ars)).alignment = right
                    group_subtotal_owner += expense.amount_ars
                    owner_total += expense.amount_ars
                    for i, _m in enumerate(members, start=4):
                        ws.cell(row=row, column=i, value="—").alignment = center
                else:
                    ws.cell(row=row, column=3, value="—").alignment = center
                    for i, m in enumerate(members, start=4):
                        if expense.paid_by_id == m.pk:
                            ws.cell(
                                row=row, column=i, value=float(expense.amount_ars)
                            ).alignment = right
                            group_subtotal_members[m.pk] += expense.amount_ars
                            member_totals[m.pk] += expense.amount_ars
                        else:
                            ws.cell(row=row, column=i, value="—").alignment = center
                row += 1

            # Subtotal del grupo
            ws.cell(row=row, column=1, value=f"Subtotal {group.name}").font = Font(italic=True)
            for col in range(1, len(col_headers) + 1):
                ws.cell(row=row, column=col).fill = subtotal_fill

            owner_cell = ws.cell(row=row, column=3)
            owner_cell.value = float(group_subtotal_owner) if group_subtotal_owner else "—"
            owner_cell.alignment = right

            for i, m in enumerate(members, start=4):
                val = group_subtotal_members.get(m.pk, 0)
                cell = ws.cell(row=row, column=i)
                cell.value = float(val) if val else "—"
                cell.alignment = right
            row += 1

        # Fila de total general
        ws.cell(row=row, column=1, value=f"TOTAL {month_name.upper()} {year}").font = total_font
        for col in range(1, len(col_headers) + 1):
            ws.cell(row=row, column=col).fill = total_fill

        ws.cell(row=row, column=3, value=float(owner_total)).font = total_font
        ws.cell(row=row, column=3).alignment = right
        for i, m in enumerate(members, start=4):
            cell = ws.cell(row=row, column=i)
            cell.value = float(member_totals.get(m.pk, 0))
            cell.font = total_font
            cell.alignment = right

        # Anchos de columna
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 12
        for col in range(3, len(col_headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 16

        # Respuesta HTTP
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"gastos_compartidos_{month_name.lower()}_{year}.xlsx"
        response = HttpResponse(
            output.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response
